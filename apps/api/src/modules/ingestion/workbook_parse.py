"""Parse restaurant finance workbook (Бистро / Benedict style) into metric rows.

Sheets map to source systems (logical origin in one Excel file):
- Отчет о фин.рез. / Расходы / баланс → 1C
- Деление БарКухня → RKeeper
- Аналитическая форма → Storyhouse / analytics
CRM lead rows are not present in this workbook shape.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

MONTHS_RU = (
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
)

KEY_PNL_NAMES = {
    "выручка, всего, в т.ч.",
    "количество гостей/заказов:",
    "средний чек",
    "себестоимость итого",
    "валовая прибыль",
    "операционные расходы",
    "операционная прибыль",
    "чистая прибыль",
    "фудкост",
}


def _slug(text: str) -> str:
    cleaned = re.sub(r"[^\wа-яА-Я0-9]+", "_", text.strip().lower(), flags=re.UNICODE)
    return cleaned.strip("_")[:80] or "metric"


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(" ", "").replace(",", ".")
        if not text or text in {"-", "—"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _month_columns(headers: list[Any]) -> list[tuple[int, str]]:
    found: list[tuple[int, str]] = []
    for idx, cell in enumerate(headers):
        if cell is None:
            continue
        if isinstance(cell, datetime):
            found.append((idx, f"{cell.year}-{cell.month:02d}"))
            continue
        name = str(cell).strip().lower()
        if name in MONTHS_RU:
            found.append((idx, name))
    return found


def parse_bistro_workbook(data: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for workbook import") from exc

    from io import BytesIO

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    metrics: list[dict[str, Any]] = []
    sheet_summary: list[dict[str, Any]] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            origin = _sheet_origin(sheet_name)
            count_before = len(metrics)
            if "фин" in sheet_name.lower() or "выруч" in sheet_name.lower():
                metrics.extend(_parse_pnl_like(sheet_name, origin, rows, unit="тыс. руб."))
            elif "расход" in sheet_name.lower():
                metrics.extend(_parse_expenses(sheet_name, origin, rows))
            elif "баркухн" in sheet_name.lower().replace(" ", "") or "бар" in sheet_name.lower():
                metrics.extend(_parse_bar_kitchen(sheet_name, origin, rows))
            elif "аналит" in sheet_name.lower():
                metrics.extend(_parse_analytics_hint(sheet_name, origin, rows))
            sheet_summary.append(
                {
                    "sheet": sheet_name,
                    "origin": origin,
                    "metrics_extracted": len(metrics) - count_before,
                }
            )
    finally:
        wb.close()

    # Keep pilot manageable: prefer key P&L + all bar/kitchen revenue/foodcost + top expenses
    metrics = _prioritize_metrics(metrics)
    return {
        "workbook_kind": "bistro_finance",
        "sheets": sheet_summary,
        "metrics": metrics,
        "notes": [
            "CRM-лиды в файле не найдены — это сводная финмодель (1C / RKeeper / аналитика).",
            "P&L значения интерпретируются как тыс. руб.; Бар/Кухня — в рублях.",
        ],
    }


def _sheet_origin(sheet_name: str) -> str:
    lower = sheet_name.lower()
    if "бар" in lower or "кухн" in lower:
        return "rkeeper"
    if "аналит" in lower:
        return "storyhouse"
    if "баланс" in lower or "расход" in lower or "фин" in lower:
        return "1c"
    return "workbook"


def _parse_pnl_like(
    sheet_name: str,
    origin: str,
    rows: list[tuple[Any, ...]],
    *,
    unit: str,
) -> list[dict[str, Any]]:
    headers = list(rows[0])
    months = _month_columns(headers)
    total_idx = next(
        (i for i, h in enumerate(headers) if h is not None and str(h).strip().lower() == "итого"),
        None,
    )
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        name = row[0] if row else None
        if not name or not str(name).strip():
            continue
        title = str(name).strip()
        monthly: dict[str, float] = {}
        for idx, month in months:
            if idx < len(row):
                num = _num(row[idx])
                if num is not None:
                    monthly[month] = num
        total = _num(row[total_idx]) if total_idx is not None and total_idx < len(row) else None
        if total is None and monthly:
            total = sum(monthly.values())
        if total is None and not monthly:
            continue
        row_unit = unit
        title_l = title.lower()
        if "гост" in title_l or "заказ" in title_l:
            row_unit = "шт."
        elif "средний чек" in title_l or title_l.startswith("чек"):
            row_unit = "руб."
        elif "%" in title or "фудкост" in title_l or "доля" in title_l:
            row_unit = "ratio"
        out.append(
            {
                "external_id": f"pnl:{_slug(title)}",
                "subject": title,
                "predicate": "finance_metric",
                "record_kind": "metric",
                "value": total,
                "unit": row_unit,
                "period": "2026-ytd",
                "months": monthly,
                "sheet": sheet_name,
                "system_origin": origin,
                "metric_group": "pnl",
            }
        )
    return out


def _parse_expenses(sheet_name: str, origin: str, rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    headers = list(rows[0])
    months = _month_columns(headers)
    total_idx = next(
        (i for i, h in enumerate(headers) if h is not None and str(h).strip().lower() == "итого"),
        None,
    )
    code_idx = next(
        (i for i, h in enumerate(headers) if h is not None and str(h).strip().lower() == "код"),
        None,
    )
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        name = row[0] if row else None
        if not name or not str(name).strip():
            continue
        title = str(name).strip()
        total = _num(row[total_idx]) if total_idx is not None and total_idx < len(row) else None
        if total is None or total == 0:
            # keep hierarchy parents with zero only if code present — skip zeros for noise
            continue
        monthly: dict[str, float] = {}
        for idx, month in months:
            if idx < len(row):
                num = _num(row[idx])
                if num is not None:
                    monthly[month] = num
        code = None
        if code_idx is not None and code_idx < len(row) and row[code_idx]:
            code = str(row[code_idx]).strip()
        out.append(
            {
                "external_id": f"exp:{code or _slug(title)}",
                "subject": title,
                "predicate": "expense_article",
                "record_kind": "metric",
                "value": total,
                "unit": "руб.",
                "period": "2026-ytd",
                "months": monthly,
                "article_code": code,
                "sheet": sheet_name,
                "system_origin": origin,
                "metric_group": "expenses",
            }
        )
    return out


def _parse_bar_kitchen(sheet_name: str, origin: str, rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    headers = list(rows[0])
    months = _month_columns(headers)
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        venue = str(row[0]).strip()
        metric = str(row[1]).strip() if row[1] else ""
        channel = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        division = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        if not metric or metric.lower().startswith("провер"):
            continue
        monthly: dict[str, float] = {}
        for idx, month in months:
            if idx < len(row):
                num = _num(row[idx])
                if num is not None:
                    monthly[month] = num
        if not monthly:
            continue
        metric_l = metric.lower()
        is_ratio = "фудкост" in metric_l or "%" in metric_l
        # ratios must not be summed across months
        total = (sum(monthly.values()) / len(monthly)) if is_ratio else sum(monthly.values())
        label = " / ".join(p for p in (venue, metric, channel, division) if p)
        out.append(
            {
                "external_id": f"rk:{_slug(label)}",
                "subject": label,
                "predicate": "ops_metric",
                "record_kind": "metric",
                "value": total,
                "unit": "ratio" if is_ratio else "руб.",
                "period": "2026-ytd",
                "months": monthly,
                "sheet": sheet_name,
                "system_origin": origin,
                "metric_group": "rkeeper",
                "venue": venue,
                "metric_name": metric,
                "channel": channel or None,
                "division": division or None,
            }
        )
    return out


def _parse_analytics_hint(sheet_name: str, origin: str, rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    # Analytics sheet is wide/complex; extract a presence marker for Storyhouse lineage.
    return [
        {
            "external_id": "storyhouse:analytics_present",
            "subject": "Аналитическая форма (Storyhouse)",
            "predicate": "workbook_marker",
            "record_kind": "metric",
            "value": 1,
            "unit": "flag",
            "period": "2026",
            "months": {},
            "sheet": sheet_name,
            "system_origin": origin,
            "metric_group": "storyhouse",
        }
    ]


def _prioritize_metrics(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    key_pnl: list[dict[str, Any]] = []
    other_pnl: list[dict[str, Any]] = []
    expenses: list[dict[str, Any]] = []
    rkeeper: list[dict[str, Any]] = []
    other: list[dict[str, Any]] = []
    for m in metrics:
        group = m.get("metric_group")
        subject = str(m.get("subject") or "").strip().lower()
        if group == "pnl":
            if subject in KEY_PNL_NAMES or any(k in subject for k in ("выручка", "чистая прибыль", "средний чек", "гостей")):
                key_pnl.append(m)
            else:
                other_pnl.append(m)
        elif group == "expenses":
            expenses.append(m)
        elif group == "rkeeper":
            rkeeper.append(m)
        else:
            other.append(m)
    expenses_sorted = sorted(expenses, key=lambda x: float(x.get("value") or 0), reverse=True)
    return key_pnl + rkeeper + expenses_sorted[:40] + other + other_pnl[:20]
