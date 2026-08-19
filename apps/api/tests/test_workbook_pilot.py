"""Workbook / bistro pilot import."""

from io import BytesIO

import pytest


def _mini_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет о фин.рез."
    ws.append(
        [
            "Наименование статей",
            "январь",
            "февраль",
            "март",
            "апрель",
            "май",
            "июнь",
            "Итого",
        ]
    )
    ws.append(["Выручка, всего, в т.ч.", 0, 0, 100, 200, 300, 400, 1000])
    ws.append(["Чистая прибыль", 0, 0, 10, 20, 30, 40, 100])
    ws.append(["Средний чек", 0, 0, 1500, 1600, 1700, 1550, 1587])

    ws2 = wb.create_sheet("Деление БарКухня")
    ws2.append([None, None, None, None, None, None, "март", "апрель", "май", "июнь"])
    ws2.append(["Бенедикт", "Выручка", "Зал", "Кухня", None, None, 1000, 2000, 3000, 4000])
    ws2.append(["Бенедикт", "Фудкост", None, "Кухня", None, None, 0.3, 0.29, 0.3, 0.3])

    ws3 = wb.create_sheet("Расходы")
    ws3.append(
        ["Статьи расходов", "январь", "февраль", "март", "апрель", "май", "июнь", "Итого", "Код"]
    )
    ws3.append(["Фонд оплаты труда", 0, 0, 1, 1, 1, 1, 5000, "00-000174"])

    ws4 = wb.create_sheet("Аналитическая форма")
    ws4.append([None, "Наименование статей"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_bistro_workbook_metrics():
    from modules.ingestion.workbook_parse import parse_bistro_workbook

    parsed = parse_bistro_workbook(_mini_workbook_bytes())
    assert parsed["workbook_kind"] == "bistro_finance"
    assert parsed["metrics"]
    origins = {m["system_origin"] for m in parsed["metrics"]}
    assert "1c" in origins
    assert "rkeeper" in origins
    assert any("Выручка" in m["subject"] for m in parsed["metrics"])


@pytest.mark.asyncio
async def test_bistro_pilot_run(client):
    data = _mini_workbook_bytes()
    response = await client.post(
        "/api/v1/pilot/bistro/run",
        files={
            "file": (
                "bistro.xlsx",
                data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["company_id"]
    assert body["analysis_id"]
    assert body["analysis_blocked"] is False
    assert body["import"]["fact_count"] >= 1
    assert body["knowledge_ids"]
    analysis = await client.get(f"/api/v1/analyses/{body['analysis_id']}")
    assert analysis.status_code == 200
    payload = analysis.json()
    assert payload["blocked"] is False
    assert payload["output"]["recommendations"] or payload["recommendations"] is not None
    briefing = (payload.get("output") or {}).get("finance_briefing")
    assert briefing, "finance analysis must include concrete briefing"
    assert briefing.get("meanings")
    assert briefing.get("summary")
    assert briefing.get("actions")
    assert body.get("conclusions"), "pilot must return conclusions for the executive UI"
    assert body["conclusions"].get("summary")
    assert body["conclusions"].get("top_expenses") is not None
    assert body["conclusions"].get("money_leaks") is not None
    assert body["conclusions"].get("profitability") is not None
    assert body.get("finance_briefing")


@pytest.mark.asyncio
async def test_bistro_pilot_ignores_stale_company_id(client):
    data = _mini_workbook_bytes()
    stale = "00000000-0000-4000-8000-000000000099"
    response = await client.post(
        "/api/v1/pilot/bistro/run",
        data={"company_id": stale},
        files={
            "file": (
                "bistro.xlsx",
                data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["company_id"] != stale
    assert body["analysis_blocked"] is False